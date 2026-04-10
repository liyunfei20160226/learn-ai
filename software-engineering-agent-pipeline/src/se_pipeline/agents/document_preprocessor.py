"""
文档预处理Agent - 处理用户提供的项目参考资料目录
流程：
1. 递归遍历目录下所有文件
2. 根据文件类型选择处理方式：
   - 纯文本：直接读取 → LLM总结
   - Office/其他：LibreOffice转PDF → 分页转图片 → 视觉模型逐页提取文本 → LLM总结
   - 原生PDF：直接分页转图片 → 视觉模型提取 → 总结
   - 不支持：跳过记录错误
3. 所有文档总结拼接成 project_background，供后续需求分析师使用
"""
import base64
from pathlib import Path
from typing import List, Tuple
from langchain_openai import ChatOpenAI
from langchain_core.messages import HumanMessage

from .base import BaseAgent
from ..types.pipeline import PipelineState, AttachedDocument
from ..prompts import get_prompt
from ..document_converter import (
    LibreOfficeConverter,
    PDFToImageConverter,
    PDFTextExtractor,
)


# 可以直接读取的纯文本扩展名
TEXT_EXTENSIONS = {
    ".txt", ".md", ".markdown",
    ".yaml", ".yml", ".json", ".jsonl",
    ".py", ".js", ".ts", ".java", ".c", ".cpp", ".h",
    ".html", ".css", ".scss", ".jsx", ".tsx",
    ".go", ".rs", ".rb", ".php", ".sh", ".bash",
    ".csv", ".sql", ".xml", ".ini", ".toml", ".cfg",
}

# 需要转换的二进制扩展名
NEED_CONVERSION_EXTENSIONS = {
    ".docx", ".doc",
    ".xlsx", ".xls",
    ".pptx", ".ppt",
    ".odt", ".ods", ".odp",
    ".rtf",
}


class DocumentPreprocessorAgent(BaseAgent):
    """文档预处理Agent"""

    def __init__(self, llm: ChatOpenAI, vision_llm: ChatOpenAI = None):
        """
        :param llm: 文本LLM，用于总结文档
        :param vision_llm: 多模态视觉LLM，用于图片文字提取，如果None用同一个llm
        """
        self.llm = llm
        self.vision_llm = vision_llm or llm
        self.libreoffice = LibreOfficeConverter()
        self.pdf_to_images = PDFToImageConverter(dpi=300)
        self.pdf_extractor = PDFTextExtractor()

    def get_output_dir(self, project_dir: Path) -> Path:
        """获取处理输出目录"""
        out_dir = project_dir / "processed_docs"
        out_dir.mkdir(parents=True, exist_ok=True)
        return out_dir

    def process_directory(
        self,
        source_dir: Path,
        output_dir: Path,
        project_id: str,
        progress_callback: callable = None
    ) -> Tuple[List[AttachedDocument], str]:
        """递归处理目录下所有文件
        :param progress_callback: 进度回调 callback(current_file: str) -> None
        :return: (文档列表，拼接后的项目背景总结)
        """
        documents: List[AttachedDocument] = []
        summaries: List[str] = []

        # 递归遍历所有文件
        file_count = 0
        for file_path in source_dir.rglob("*"):
            if file_path.is_dir():
                continue
            # 跳过处理输出目录，避免重复扫描转换后的文件
            if "processed_docs" in str(file_path):
                continue
            # 跳过隐藏文件和临时文件
            if file_path.name.startswith('.') or file_path.name.startswith('~$'):
                continue
            # 跳过我们生成的汇总文件
            if file_path.name in ["pipeline_state.yaml", "01-requirements.yaml", "01-requirements-spec.md", "qa-history.yaml"]:
                continue

            file_count += 1
            rel_path = str(file_path.relative_to(source_dir))

            # 通知进度
            if progress_callback:
                progress_callback(rel_path)

            print(f"  正在处理: {rel_path}... ", end="", flush=True)
            if progress_callback:
                progress_callback(f"正在处理: {rel_path}")

            doc = self.process_single_file(file_path, source_dir, output_dir, progress_callback)
            documents.append(doc)

            if doc.parse_success:
                # 从summary文件读取内容
                if doc.summary_path:
                    summary_path = output_dir / doc.summary_path
                    with open(summary_path, "r", encoding="utf-8") as f:
                        summary_content = f.read()
                    summaries.append(f"## 文档: {doc.relative_path}\n\n{summary_content}\n\n")
                    # 统计字符数（summary_path相对路径长度不重要，看实际summary内容）
                    print(f"✓ 成功 ({len(summary_content):,d} 字符)")
            else:
                print(f"✗ 跳过 - {doc.error_message}")

        if file_count > 0:
            print()

        # 拼接所有总结
        if summaries:
            individual_summaries = "".join(summaries)
            # 再做一次整体汇总，消除重复，整合信息
            full_background = self._final_summary(individual_summaries)
        else:
            full_background = ""

        return documents, full_background

    def process_single_file(
        self,
        file_path: Path,
        source_dir: Path,
        output_dir: Path,
        progress_callback: callable = None
    ) -> AttachedDocument:
        """处理单个文件"""
        relative_path = str(file_path.relative_to(source_dir))
        ext = file_path.suffix.lower()

        doc = AttachedDocument(
            filename=file_path.name,
            relative_path=relative_path,
            absolute_path=str(file_path.absolute()),
            original_ext=ext,
            file_size=file_path.stat().st_size,
        )

        try:
            if ext in TEXT_EXTENSIONS:
                # 纯文本直接读取
                self._process_text_file(doc, file_path, output_dir)
            elif ext in ['.xlsx', '.xls', '.xlsm']:
                # Excel原生解析 - pandas读表格转markdown，比PDF转图片更准确
                self._process_excel(doc, file_path, output_dir, progress_callback)
            elif ext == '.pdf':
                # 原生PDF，不需要转换，直接处理
                self._process_pdf(doc, file_path, output_dir, progress_callback)
            elif ext in NEED_CONVERSION_EXTENSIONS and self.libreoffice.can_convert(file_path):
                # 需要用LibreOffice转换为PDF
                self._process_with_conversion(doc, file_path, output_dir, progress_callback)
            elif ext in {'.png', '.jpg', '.jpeg', '.gif', '.bmp', '.tiff', '.tif'}:
                # 图片，直接用视觉模型解析
                self._process_image(doc, file_path, output_dir)
            else:
                # 不支持的格式
                doc.parse_success = False
                doc.error_message = f"Unsupported file format: {ext}"

        except Exception as e:
            doc.parse_success = False
            doc.error_message = f"Processing exception: {str(e)}"

        return doc

    def _process_text_file(self, doc: AttachedDocument, file_path: Path, output_dir: Path) -> None:
        """处理纯文本文件"""
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                content = f.read()
            # LLM总结
            summary = self._summarize_content(doc.filename, doc.original_ext, content)
            # 保存summary到单独文件，只存路径
            summary_filename = f"{file_path.stem}.summary.md"
            summary_path = output_dir / summary_filename
            with open(summary_path, "w", encoding="utf-8") as f:
                f.write(f"# 文档: {doc.relative_path}\n\n{summary}\n")
            doc.summary_path = str(summary_path.relative_to(output_dir))
            doc.parse_success = True
        except UnicodeDecodeError:
            # 不是纯文本，可能是二进制
            doc.parse_success = False
            doc.error_message = "File is not valid UTF-8 text"

    def _process_with_conversion(
        self,
        doc: AttachedDocument,
        input_path: Path,
        output_dir: Path,
        progress_callback: callable = None
    ) -> None:
        """先用LibreOffice转换为PDF，再处理PDF"""
        # 为当前文件创建单独子目录
        file_subdir = output_dir / input_path.stem
        file_subdir.mkdir(parents=True, exist_ok=True)
        # 转换PDF放到子目录
        result = self.libreoffice.convert(input_path, file_subdir)
        if not result.success or not result.output_path:
            doc.parse_success = False
            doc.error_message = result.error_message or "Conversion failed"
            return

        doc.processed_pdf_path = str(result.output_path.relative_to(output_dir))
        self._process_pdf(doc, result.output_path, output_dir, progress_callback)


    def _process_excel(
        self,
        doc: AttachedDocument,
        excel_path: Path,
        output_dir: Path,
        progress_callback: callable = None
    ) -> None:
        """处理Excel文件原生解析 - 每个sheet单独处理
        改进方案（双路径提取 + LLM合并）：
        1. 第一步：pandas读取每个sheet，转dataframe → 转markdown表格
           → 保证表格数据准确，但浮动图片可能抓不到
        2. 第二步：整个Excel转PDF → 逐页转图片 → 多模态逐页OCR提取文字
           → 能抓到页面上所有内容，包括浮动图片/图表，但表格格式可能不准
        3. 第三步：LLM合并两份结果 → 保留准确表格，插入PDF提取的图片内容
           → 解决图片无法归属的问题

        目录结构：
        processed_docs/
          ├─ {original_filename}/             <- 每个文件单独子目录
          │   ├─ pdf_pages/                  <- PDF转图片存放这里
          │   ├─ sheet_xx_name.summary.md    <- 每个sheet单独总结
          │   └─ full_merged_content.md      <- LLM合并后的完整内容
          └─ {original_filename}.summary.md  <- 文件总summary在根目录
        """
        try:
            import openpyxl
            import pandas as pd
        except ImportError:
            doc.parse_success = False
            doc.error_message = "Missing dependencies: need openpyxl and pandas. Install with: pip install openpyxl pandas"
            return

        wb = openpyxl.load_workbook(excel_path, data_only=True)

        # ========== 为当前Excel创建单独子目录 ==========
        file_subdir = output_dir / excel_path.stem
        file_subdir.mkdir(parents=True, exist_ok=True)

        # ========== 第一步：遍历每个sheet，读取表格数据转markdown ==========
        # 存储每个sheet的信息：(sheet_idx, sheet_name, content_parts, extracted_text)
        sheets_info: list[tuple[int, str, list[str], str]] = []

        for sheet_idx, sheet_name in enumerate(wb.sheetnames, 1):
            ws = wb[sheet_name]
            print(f"    Processing sheet: {sheet_name}... ", end="", flush=True)
            if progress_callback:
                progress_callback(f"{doc.filename} → 读取 sheet {sheet_idx}/{len(wb.sheetnames)}: {sheet_name}")

            # 读取表格数据转markdown
            data = []
            for row in ws.iter_rows(values_only=True):
                data.append(list(row))

            sheet_markdown = ""
            if data and any(cell is not None for row in data for cell in row):
                df = pd.DataFrame(data)
                # 去掉全空行
                df = df.dropna(how='all')
                if not df.empty:
                    sheet_markdown = df.to_markdown(index=False, tablefmt="github")

            # 构建内容（只有表格文字，图片由PDF OCR提取，后续LLM合并）
            content_parts: list[str] = []
            content_parts.append(f"# Sheet: {sheet_name}\n")
            if sheet_markdown:
                content_parts.append(sheet_markdown)
                content_parts.append("")

            extracted_text = "\n".join(content_parts)

            # 保存sheet信息
            sheets_info.append((sheet_idx, sheet_name, content_parts, extracted_text))
            print("✓ 表格读取完成")

        # ========== 逐个sheet总结，保存单独文件 ==========
        all_sheet_summaries: list[str] = []
        structured_text_parts: list[str] = []

        for (sheet_idx, sheet_name, content_parts, extracted_text) in sheets_info:
            print(f"    正在总结: Sheet {sheet_idx}: {sheet_name}... ", end="", flush=True)
            if progress_callback:
                progress_callback(f"{doc.filename} → 总结 sheet {sheet_idx}/{len(sheets_info)}: {sheet_name}")
            # 总结这个sheet
            sheet_summary = self._summarize_content(
                f"{doc.filename} - Sheet {sheet_idx}: {sheet_name}",
                "excel sheet",
                extracted_text
            )

            # 保存sheet单独summary - 放到文件子目录
            sheet_summary_filename = f"sheet_{sheet_idx:02d}_{sheet_name}.summary.md"
            sheet_summary_path = file_subdir / sheet_summary_filename
            with open(sheet_summary_path, "w", encoding="utf-8") as f:
                f.write(f"# {doc.filename} - Sheet {sheet_idx}: {sheet_name}\n\n{sheet_summary}\n")

            # 收集结构化文本（每个sheet原始提取结果）
            all_sheet_summaries.append(f"## Sheet {sheet_idx}: {sheet_name}\n\n{sheet_summary}\n")
            structured_text_parts.append(f"--- Sheet {sheet_idx}: {sheet_name} ---\n{extracted_text}\n")
            # sheet 之间加延迟，避免触发频率限制
            import time
            time.sleep(0.1)
            print(f"✓ 完成 ({len(sheet_summary)} 字符)")

        wb.close()

        # ========== 第二步：将整个Excel转PDF，逐页OCR提取文字 ==========
        pdf_pages_text = self._convert_excel_to_pdf_and_ocr(doc, excel_path, file_subdir, progress_callback)

        # ========== 第三步：合并结构化表格结果 + PDF OCR结果，LLM生成完整文本 ==========
        # 合并结果写入单独文件，供后续读取
        if pdf_pages_text:
            # 构建完整文本送给LLM合并
            structured_text = "\n\n".join(structured_text_parts)
            if progress_callback:
                progress_callback(f"{doc.filename} → LLM合并结果中...")
            merged_text = self._merge_excel_results(structured_text, pdf_pages_text)
            # 保存完整合并结果到单独文件
            full_merged_path = file_subdir / "full_merged_content.md"
            with open(full_merged_path, "w", encoding="utf-8") as f:
                f.write(merged_text)
            print(f"    ✓ 完整合并结果保存完成: {len(merged_text)} 字符")
            if progress_callback:
                progress_callback(f"{doc.filename} → ✓ 完整合并结果保存完成")

        # ========== 汇总保存 ==========
        if all_sheet_summaries:
            full_summary = f"# {doc.filename}\n\n" + "\n".join(all_sheet_summaries)
            # 保存summary到单独文件，只存路径
            summary_filename = f"{excel_path.stem}.summary.md"
            summary_path = output_dir / summary_filename
            with open(summary_path, "w", encoding="utf-8") as f:
                f.write(full_summary)
            doc.summary_path = str(summary_path.relative_to(output_dir))
            doc.parse_success = True
        else:
            doc.parse_success = False
            doc.error_message = "No data extracted from any sheet"

    def _process_pdf(
        self,
        doc: AttachedDocument,
        pdf_path: Path,
        output_dir: Path,
        progress_callback: callable = None
    ) -> None:
        """处理PDF：转图片，视觉提取，总结

        统一策略：
        - 只有原始文件就是纯文本格式 → 直接读取文本，不走图片
        - 任何其他情况（PDF, Excel, Word, 图片 等）→ 分页转图片，多模态逐页提取，逐页保存

        目录结构：
        processed_docs/
          ├─ {original_filename}/     <- 每个文件单独子目录
          │   ├─ images/              <- 所有分页图片放这里
          │   ├─ {pdf}_page_001.summary.md
          │   ├─ {pdf}_page_002.summary.md
          │   ...
          └─ {original_filename}.summary.md  <- 文件总summary在根目录
        """
        # 获取PDF页数
        success_extract, text, page_count = self.pdf_extractor.extract(pdf_path)

        full_text_parts: List[str] = []
        page_summaries: List[str] = []
        use_per_page_summary = False

        # 为当前PDF创建单独子目录
        file_subdir = output_dir / pdf_path.stem
        file_subdir.mkdir(parents=True, exist_ok=True)
        images_dir = file_subdir / "images"
        images_dir.mkdir(exist_ok=True)

        # 判断：只有原始纯文本文件才走直接提取
        is_original_text = doc.original_ext in TEXT_EXTENSIONS

        if is_original_text and success_extract and text.strip() and len(text.strip()) > 100:
            # 原始纯文本 → 直接用提取的文本
            full_text_parts.append(text)
        else:
            # 所有非纯文本原始文件 → 强制转图片逐页处理（多模态对表格排版理解更好）
            image_result = self.pdf_to_images.convert(pdf_path, file_subdir / "images")
            if not image_result.success:
                doc.parse_success = False
                doc.error_message = image_result.error_message or "PDF to image failed"
                return

            # 更新图片路径为相对路径
            doc.page_image_paths = [str(p.relative_to(output_dir)) for p in image_result.image_paths]
            # 逐页视觉提取 + 逐页保存summary

            for page_num, img_path in enumerate(image_result.image_paths, 1):
                if progress_callback:
                    progress_callback(f"{doc.filename} → 提取第{page_num}页")
                page_text = self._extract_text_from_image(img_path)
                full_text_parts.append(f"--- 第{page_num}页 ---\n{page_text}")

                if progress_callback:
                    progress_callback(f"{doc.filename} → 总结第{page_num}页")
                # 每页都保存单独summary文件，方便人工review
                page_summary = self._summarize_content(
                    f"{doc.filename} - 第{page_num}页",
                    f"page {page_num} of {doc.original_ext}",
                    page_text
                )
                page_summaries.append(f"## 第{page_num}页\n\n{page_summary}\n")

                # 单独保存每页summary文件，放到子目录
                summary_filename = f"page_{page_num:03d}.summary.md"
                summary_path = file_subdir / summary_filename
                with open(summary_path, "w", encoding="utf-8") as f:
                    f.write(f"# {doc.filename} - 第{page_num}页\n\n{page_summary}\n")

            use_per_page_summary = len(image_result.image_paths) >= 1 and len(page_summaries) > 0

        # LLM总结
        if full_text_parts:
            if use_per_page_summary and page_summaries:
                # 多页，已经逐页总结，直接拼接
                full_summary = "# {}\n\n{}".format(doc.filename, "\n".join(page_summaries))
            else:
                # 单页，整体总结一次
                full_text = "\n\n".join(full_text_parts)
                full_summary = self._summarize_content(
                    doc.filename,
                    doc.original_ext,
                    full_text
                )
            # 保存summary到根目录，只存路径
            summary_filename = f"{pdf_path.stem}.summary.md"
            summary_path = output_dir / summary_filename
            with open(summary_path, "w", encoding="utf-8") as f:
                f.write(full_summary)
            doc.summary_path = str(summary_path.relative_to(output_dir))
            doc.parse_success = True
        else:
            doc.parse_success = False
            doc.error_message = "No text extracted from PDF"

    def _process_image(
        self,
        doc: AttachedDocument,
        image_path: Path,
        output_dir: Path
    ) -> None:
        """处理单个图片文件，直接用视觉模型提取文字

        目录结构：
        processed_docs/
          ├─ {original_filename}/     <- 每个文件单独子目录
          │   ├─ images/              <- 原始图片放这里
          └─ {original_filename}.summary.md  <- 文件总summary在根目录
        """
        # 为当前图片创建单独子目录
        file_subdir = output_dir / image_path.stem
        file_subdir.mkdir(parents=True, exist_ok=True)
        images_dir = file_subdir / "images"
        images_dir.mkdir(exist_ok=True)

        # 复制原始图片到子目录
        output_img_path = images_dir / image_path.name
        if not output_img_path.exists():
            import shutil
            shutil.copy2(image_path, output_img_path)

        doc.page_image_paths = [str(output_img_path.relative_to(output_dir))]
        extracted_text = self._extract_text_from_image(output_img_path)

        if extracted_text.strip():
            summary = self._summarize_content(doc.filename, doc.original_ext, extracted_text)
            # 保存summary到根目录，只存路径
            summary_filename = f"{image_path.stem}.summary.md"
            summary_path = output_dir / summary_filename
            with open(summary_path, "w", encoding="utf-8") as f:
                f.write(f"# 文档: {doc.relative_path}\n\n{summary}\n")
            doc.summary_path = str(summary_path.relative_to(output_dir))
            doc.parse_success = True
        else:
            doc.parse_success = False
            doc.error_message = "No text extracted from image"

    def _extract_text_from_image(self, image_path: str) -> str:
        """用视觉模型从图片提取文字"""
        import time
        prompt_template = get_prompt("document_vision_parser")

        # 读取图片并编码
        with open(image_path, "rb") as f:
            image_bytes = f.read()
        base64_image = base64.b64encode(image_bytes).decode('utf-8')

        # 构造多模态消息
        message = HumanMessage(
            content=[
                {"type": "text", "text": prompt_template},
                {
                    "type": "image_url",
                    "image_url": {
                        "url": f"data:image/png;base64,{base64_image}"
                    }
                },
            ]
        )

        response = self.vision_llm.invoke([message])
        # 小延迟避免触发频率限制 429
        time.sleep(0.2)
        return response.content.strip()

    def _summarize_content(self, filename: str, file_type: str, content: str) -> str:
        """用LLM总结文档内容"""
        # 如果内容很短，不需要总结，直接返回
        if len(content) < 500:
            return content

        prompt_template = get_prompt("document_preprocessor")
        prompt = prompt_template\
            .replace("{{FILE_PATH}}", filename)\
            .replace("{{FILE_TYPE}}", file_type)\
            .replace("{{DOC_CONTENT}}", content)

        import time
        response = self.llm.invoke(prompt)
        # 小延迟避免触发频率限制 429
        time.sleep(0.2)
        return response.content.strip()

    def _final_summary(self, individual_summaries: str) -> str:
        """对所有文档的单独总结做最终整合汇总，消除重复"""
        import time
        prompt_template = get_prompt("document_final_summary")
        prompt = prompt_template.replace("{{DOC_SUMMARIES}}", individual_summaries)
        response = self.llm.invoke(prompt)
        final_summary = response.content.strip()
        # 小延迟避免触发频率限制 429
        time.sleep(0.2)
        # 添加标题
        if not final_summary.startswith("#"):
            final_summary = "# 项目参考资料汇总\n\n" + final_summary
        return final_summary

    def run(self, state: PipelineState, progress_callback: callable = None) -> PipelineState:
        """执行文档预处理
        :param progress_callback: 进度回调 callback(current_file: str) -> None
        """
        if not state.source_documents_dir:
            # 没有资料目录，直接返回
            state.documents_processed = True
            return state

        # 尝试解析路径
        source_dir = Path(state.source_documents_dir)
        if not source_dir.is_absolute():
            # 相对路径，先尝试相对于当前工作目录
            cwd = Path.cwd()
            source_dir_cwd = cwd / state.source_documents_dir
            if source_dir_cwd.exists():
                source_dir = source_dir_cwd
            else:
                # 尝试相对于项目根目录（当前文件的三级上级）
                root = Path(__file__).parent.parent.parent.parent
                source_dir_root = root / state.source_documents_dir
                if source_dir_root.exists():
                    source_dir = source_dir_root

        if not source_dir.exists():
            print(f"⚠️  资料目录不存在: {source_dir}")
            print("   请检查路径是否正确，建议使用绝对路径")
            state.documents_processed = False
            return state

        print(f"🔍 开始扫描目录: {source_dir.absolute()}")

        # 获取项目输出目录
        from ..storage.project_store import ProjectStore
        store = ProjectStore()
        project_dir = store.get_project_dir(state.project_id)
        output_dir = self.get_output_dir(project_dir)

        # 处理所有文件
        documents, project_background = self.process_directory(
            source_dir,
            output_dir,
            state.project_id,
            progress_callback
        )

        # 更新状态
        state.attached_documents = documents
        # project_background 只存文件名，不存完整内容
        # 空表示没有背景资料，非空表示背景资料在这个文件中
        if project_background and store:
            project_dir = store.get_project_dir(state.project_id)
            md_file = project_dir / "00-project-background.md"
            with open(md_file, "w", encoding="utf-8") as f:
                f.write(project_background)
            state.project_background = "00-project-background.md"
        else:
            state.project_background = ""
        state.documents_processed = True
        state.update_timestamp()

        return state

    def _convert_excel_to_pdf_and_ocr(
        self,
        doc: AttachedDocument,
        excel_path: Path,
        file_subdir: Path,
        progress_callback: callable = None
    ) -> list[str]:
        """第二步：将整个Excel转PDF，然后PDF逐页转图片，多模态逐页OCR提取文字

        Returns:
            list[str]: 每页提取的文字列表
        """
        # 第一步：用LibreOffice将Excel转PDF
        # LibreOffice.convert 需要传入 output_dir，它自动生成 input.stem.pdf
        if progress_callback:
            progress_callback(f"{doc.filename} → Excel转PDF")
        result = self.libreoffice.convert(excel_path, file_subdir)
        if not result.success:
            print(f"    ✗ Excel转PDF失败: {result.error_message}")
            print("    回退到只使用结构化结果")
            doc.error_message = "Excel to PDF conversion failed"
            return []

        # 获取输出的PDF路径
        pdf_path = file_subdir / f"{excel_path.stem}.pdf"
        if not pdf_path.exists() and result.output_path:
            pdf_path = result.output_path
        print(f"    ✓ Excel转PDF完成: {pdf_path.name}")

        # 第二步：PDF逐页转图片
        if progress_callback:
            progress_callback(f"{doc.filename} → PDF逐页转图片")
        image_result = self.pdf_to_images.convert(pdf_path, file_subdir / "pdf_pages")
        if not image_result.success:
            print(f"    ✗ PDF转图片失败: {image_result.error_message}")
            return []

        print(f"    ✓ PDF转图片完成: {len(image_result.image_paths)} 页")

        # 第三步：逐页OCR提取文字
        pages_text: list[str] = []
        for page_num, img_path in enumerate(image_result.image_paths, 1):
            if progress_callback:
                progress_callback(f"{doc.filename} → 第{page_num}页 OCR提取")
            page_text = self._extract_text_from_image(img_path)
            if page_text.strip():
                pages_text.append(f"--- PDF第{page_num}页 ---\n{page_text}")
            print(f"    ✓ 第{page_num}页OCR完成: {len(page_text)} 字符")
            # 小延迟避免触发频率限制
            import time
            time.sleep(0.1)

        return pages_text

    def _merge_excel_results(self, structured_text: str, pdf_pages_text: list[str]) -> str:
        """第三步：LLM合并结构化表格结果和PDF OCR结果，生成最终完整文本

        Args:
            structured_text: pandas提取的每个sheet结构化表格文字
            pdf_pages_text: PDF转图片OCR提取的每页文字

        Returns:
            str: 合并后的完整文本
        """
        if not pdf_pages_text:
            return structured_text

        pdf_combined = "\n\n".join(pdf_pages_text)

        prompt = f"""请帮我合并两份来自同一个Excel文件的提取结果，生成一份完整、准确、不重复的Excel文档内容：

# 第一份：pandas直接提取的结构化表格数据（准确，但不包含浮动图片/图表）
{structured_text[:8000]}

# 第二份：Excel转PDF后逐页OCR提取的内容（包含所有图片/图表文字，但表格可能格式不对）
{pdf_combined[:8000]}

# 任务：
请将两份结果合并，生成一份完整、不重复的最终文档：
1. 保留pandas提取的结构化表格，因为它准确
2. 将PDF OCR提取到的浮动图片、图表、文本框等内容插入到正确位置
3. 去重：相同内容只保留一份
4. 保持文档结构清晰，按sheet分页

# 输出：
请直接输出合并后的完整文档，不要额外解释。
"""
        import time
        time.sleep(0.2)
        response = self.llm.invoke(prompt)
        merged = response.content.strip()
        print(f"    ✓ LLM合并完成: {len(merged)} 字符")
        return merged

    def get_processing_stats(self, documents: List[AttachedDocument]) -> Tuple[int, int]:
        """获取处理统计"""
        total = len(documents)
        success = sum(1 for d in documents if d.parse_success)
        return total, success
