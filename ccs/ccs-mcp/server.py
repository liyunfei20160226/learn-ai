#!/usr/bin/env python3
"""
CCS MCP Server - Read CCS design documents (Excel) and execute database operations.

This MCP server provides tools to:
1. List all Excel design documents in the design directory
2. Read specific worksheet data from a design file
3. Extract embedded UI images from Excel
4. Get design file information (sheet list, image count)
5. Execute SQL statements on PostgreSQL database (CREATE TABLE, etc.)
6. List tables and describe table structure
"""

import sys
from typing import Optional, List, Dict, Any
from pathlib import Path
import os
import json

import psycopg2
from psycopg2.extras import RealDictCursor

from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import BaseModel, Field, ConfigDict
from mcp.server.fastmcp import FastMCP, Context

# Configure stdout/stderr encoding for Windows compatibility
sys.stdout.reconfigure(encoding='utf-8')
sys.stderr.reconfigure(encoding='utf-8')

# Load environment variables from .env file
load_dotenv()

# Configuration
DESIGN_DIR = Path(os.getenv("CCS_DESIGN_DIR", "D:/dev/learn-ai/ccs/设计书"))
OUTPUT_DIR = Path(os.getenv("CCS_OUTPUT_DIR", "D:/dev/learn-ai/ccs/设计书/output"))
IMAGE_OUTPUT_DIR = OUTPUT_DIR / "images"

# Database configuration
DB_HOST = os.getenv("DB_HOST", "localhost")
DB_PORT = int(os.getenv("DB_PORT", "5432"))
DB_NAME = os.getenv("DB_NAME", "ccs")
DB_USER = os.getenv("DB_USER", "postgres")
DB_PASSWORD = os.getenv("DB_PASSWORD", "postgres")

# Ensure output directories exist
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
IMAGE_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


# Initialize MCP server
mcp = FastMCP("ccs-mcp")


# Database connection helper
def _get_db_connection():
    """Get a PostgreSQL database connection."""
    return psycopg2.connect(
        host=DB_HOST,
        port=DB_PORT,
        dbname=DB_NAME,
        user=DB_USER,
        password=DB_PASSWORD
    )


# Helper functions
def _list_excel_files(design_dir: Path) -> List[Dict[str, Any]]:
    """List all Excel files in design directory."""
    excel_files = []
    for ext in [".xlsx", ".xls", ".xlsm"]:
        for file in design_dir.glob(f"*{ext}"):
            if not file.name.startswith("~$"):  # Skip temporary files
                excel_files.append({
                    "filename": file.name,
                    "path": str(file),
                    "size_kb": round(file.stat().st_size / 1024, 1)
                })
    return sorted(excel_files, key=lambda x: x["filename"])


def _convert_cell_value(value: Any) -> Any:
    """Convert cell value to JSON serializable type."""
    from datetime import datetime, date
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return value


def _read_sheet_data(file_path: Path, sheet_name: str) -> List[List[Any]]:
    """Read a single worksheet data as table."""
    wb: Workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
    ws: Worksheet = wb[sheet_name]
    data = []
    for row in ws.iter_rows(values_only=True):
        converted_row = [_convert_cell_value(cell) for cell in row]
        data.append(converted_row)
    wb.close()
    return data


def _extract_images_from_file(file_path: Path, image_output_dir: Path, force_sheet_name: Optional[str] = None) -> List[Dict[str, Any]]:
    """Extract all embedded PNG images from Excel file, organized by sheet.

    - Only create directory for sheet that actually has images
    - Put images directly in the sheet directory they belong to
    - Unknown sheet images go to 'unknown' directory (or to force_sheet_name if specified)
    - Only PNG format is extracted
    """
    extracted_images = []
    base_name = file_path.stem

    # First: Try openpyxl to extract images with sheet information
    wb: Workbook = load_workbook(filename=file_path, read_only=False, keep_vba=True)
    image_index = 1

    # Check each sheet for images
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_has_images = False

        if hasattr(ws, '_images') and ws._images:
            for image in ws._images:
                try:
                    image_data = None

                    # Get image data from various locations
                    if hasattr(image, 'ref') and hasattr(image.ref, 'blip'):
                        blip = image.ref.blip
                        if blip is not None:
                            if hasattr(blip, 'compressed_picture'):
                                image_data = blip.compressed_picture
                            elif hasattr(blip, '_blip'):
                                image_data = blip._blip
                            elif hasattr(blip, 'embed'):
                                # For some versions of openpyxl, binary data is here
                                image_data = blip.embed
                    if image_data is None and hasattr(image, 'compressed_picture'):
                        image_data = image.compressed_picture
                    if image_data is None and hasattr(image, '_blip'):
                        image_data = image._blip
                    if image_data is None and hasattr(image, '_image'):
                        image_data = image._image
                    if image_data is None and hasattr(image, 'picture'):
                        image_data = image.picture

                    if not image_data:
                        print(f"Debug: No image_data found for image in {sheet_name}", file=sys.stderr)
                        continue

                    # Check if this is actually a PNG by checking the PNG signature
                    # PNG magic number: 89 50 4E 47 0D 0A 1A 0A
                    is_png = False
                    if len(image_data) >= 8 and image_data[:8] == b'\x89PNG\r\n\x1a\n':
                        is_png = True

                    if not is_png:
                        # Check if it's EMF or other format - skip according to requirement
                        print(f"Info: Skipping non-PNG image in {sheet_name}", file=sys.stderr)
                        continue

                    # Only create directory if we actually have an image
                    if not sheet_has_images:
                        sheet_image_dir = image_output_dir / base_name / sheet_name
                        sheet_image_dir.mkdir(parents=True, exist_ok=True)
                        sheet_has_images = True
                    else:
                        sheet_image_dir = image_output_dir / base_name / sheet_name

                    image_filename = f"{base_name}_{sheet_name}_{image_index:03d}.png"
                    image_path = sheet_image_dir / image_filename

                    with open(image_path, "wb") as f:
                        f.write(image_data)

                    # Get image position
                    row = None
                    col = None
                    if hasattr(image, 'anchor'):
                        if hasattr(image.anchor, '_row'):
                            row = image.anchor._row
                        if hasattr(image.anchor, '_col'):
                            col = image.anchor._col

                    extracted_images.append({
                        "filename": image_filename,
                        "path": str(image_path),
                        "relative_path": str(sheet_image_dir.relative_to(OUTPUT_DIR) / image_filename),
                        "sheet_name": sheet_name,
                        "position": {"row": row, "col": col}
                    })
                    image_index += 1
                except Exception as e:
                    print(f"Warning: Failed to save image on {sheet_name}: {e}", file=sys.stderr)
                    continue

    wb.close()

    # Second: Check zip file for PNGs that openpyxl missed
    import zipfile
    try:
        with zipfile.ZipFile(file_path, 'r') as zf:
            for img_zip_path in zf.namelist():
                # Only extract PNG format
                if img_zip_path.startswith('xl/media/') and img_zip_path.lower().endswith('.png'):
                    # Unknown sheet images always go to 'unknown' directory
                    # Because we can't be sure which sheet they belong to
                    target_dir = image_output_dir / base_name / 'unknown'
                    target_dir.mkdir(parents=True, exist_ok=True)
                    image_filename = f"{base_name}_unknown_{image_index:03d}.png"
                    output_path = target_dir / image_filename
                    relative_path = str(target_dir.relative_to(OUTPUT_DIR) / image_filename)
                    sheet_name_result = None

                    with zf.open(img_zip_path) as source:
                        image_data = source.read()
                        with open(output_path, "wb") as dest:
                            dest.write(image_data)

                    extracted_images.append({
                        "filename": image_filename,
                        "path": str(output_path),
                        "relative_path": relative_path,
                        "sheet_name": sheet_name_result,
                        "position": {"row": None, "col": None}
                    })
                    image_index += 1
    except Exception as e:
        print(f"Warning: Zip extraction failed: {e}", file=sys.stderr)

    # Clean up: remove empty directories
    base_dir = image_output_dir / base_name
    if base_dir.exists():
        for child in base_dir.iterdir():
            if child.is_dir() and not any(child.iterdir()):
                child.rmdir()
        # Remove base dir if empty
        if not any(base_dir.iterdir()):
            base_dir.rmdir()

    return extracted_images


# Input models
class ListDesignFilesInput(BaseModel):
    """Input model for listing design files."""
    model_config = ConfigDict(
        str_strip_whitespace=True,
        validate_assignment=True,
        extra='forbid'
    )

    extension: Optional[str] = Field(
        default="xlsx",
        description="Filter by file extension: xlsx, xls, all (default: xlsx)",
        pattern=r"^(xlsx|xls|all)$"
    )


class ReadDesignSheetInput(BaseModel):
    """Input model for reading a specific worksheet."""
    model_config = ConfigDict(
        str_strip_whitespace=True,
        validate_assignment=True,
        extra='forbid'
    )

    filename: str = Field(
        ...,
        description="Filename of the Excel design document (e.g., '【CCS】xxx.xlsx')",
        min_length=1
    )
    sheet_name: str = Field(
        ...,
        description="Name of the worksheet to read (e.g., 'レイアウト', 'プログラム概要')",
        min_length=1
    )


class ExtractImagesInput(BaseModel):
    """Input model for extracting images from a design file."""
    model_config = ConfigDict(
        str_strip_whitespace=True,
        validate_assignment=True,
        extra='forbid'
    )

    filename: str = Field(
        ...,
        description="Filename of the Excel design document",
        min_length=1
    )
    sheet_name: Optional[str] = Field(
        default=None,
        description="Extract images only from this specific sheet (optional, default: all sheets)"
    )


class GetDesignInfoInput(BaseModel):
    """Input model for getting design file information."""
    model_config = ConfigDict(
        str_strip_whitespace=True,
        validate_assignment=True,
        extra='forbid'
    )

    filename: str = Field(
        ...,
        description="Filename of the Excel design document",
        min_length=1
    )


# Tool definitions

@mcp.tool(
    name="ccs_list_design_files",
    annotations={
        "title": "List All CCS Design Documents",
        "readOnlyHint": True,
        "destructiveHint": False,
        "idempotentHint": True,
        "openWorldHint": False
    }
)
async def ccs_list_design_files(params: ListDesignFilesInput, ctx: Context) -> str:
    """List all Excel design document files in the CCS design directory.

    Use this to discover which design files are available before reading them.

    Args:
        params: Validated input parameters containing:
            - extension (str): Filter by file extension: xlsx, xls, or all (default: xlsx)

    Returns:
        str: JSON formatted list of design files with their sizes
    """
    try:
        files = _list_excel_files(DESIGN_DIR)

        if params.extension != "all":
            files = [f for f in files if f["filename"].endswith(f".{params.extension}")]

        if not files:
            return json.dumps({
                "total": 0,
                "files": []
            }, indent=2, ensure_ascii=False)

        return json.dumps({
            "total": len(files),
            "design_directory": str(DESIGN_DIR),
            "files": files
        }, indent=2, ensure_ascii=False)

    except Exception as e:
        return json.dumps({
            "error": str(e)
        }, ensure_ascii=False)


@mcp.tool(
    name="ccs_read_design_sheet",
    annotations={
        "title": "Read Worksheet Data from Design File",
        "readOnlyHint": True,
        "destructiveHint": False,
        "idempotentHint": True,
        "openWorldHint": False
    }
)
async def ccs_read_design_sheet(params: ReadDesignSheetInput, ctx: Context) -> str:
    """Read the complete table data from a specific worksheet in a design file.

    This tool extracts all cell values from a worksheet and returns them as a table.
    Use this to get the design specification, layout information, business logic, etc.

    Args:
        params: Validated input parameters containing:
            - filename (str): Filename of the Excel design file (from ccs_list_design_files)
            - sheet_name (str): Name of the worksheet to read (e.g., 'レイアウト', 'プログラム概要', '処理説明')

    Returns:
        str: JSON formatted table data with all cell values
    """
    try:
        file_path = DESIGN_DIR / params.filename

        if not file_path.exists():
            return json.dumps({
                "error": f"File '{params.filename}' not found in {DESIGN_DIR}"
            }, ensure_ascii=False)

        if not file_path.is_file():
            return json.dumps({
                "error": f"'{params.filename}' is not a file"
            }, ensure_ascii=False)

        data = _read_sheet_data(file_path, params.sheet_name)

        return json.dumps({
            "filename": params.filename,
            "sheet_name": params.sheet_name,
            "rows": len(data),
            "data": data
        }, indent=2, ensure_ascii=False)

    except Exception as e:
        return json.dumps({
            "error": f"Error reading sheet: {str(e)}"
        }, ensure_ascii=False)


@mcp.tool(
    name="ccs_extract_images",
    annotations={
        "title": "Extract Embedded UI Images from Design File",
        "readOnlyHint": False,
        "destructiveHint": False,
        "idempotentHint": True,
        "openWorldHint": False
    }
)
async def ccs_extract_images(params: ExtractImagesInput, ctx: Context) -> str:
    """Extract all embedded images from a design file (usually UI screenshots/layout diagrams).

    This tool extracts all embedded images from the Excel file and saves them as PNG files
    in the output directory. It returns the paths to the extracted images.

    Args:
        params: Validated input parameters containing:
            - filename (str): Input Excel filename
            - sheet_name (optional): Extract only from this specific sheet

    Returns:
        str: JSON formatted list of extracted images with their paths
    """
    try:
        file_path = DESIGN_DIR / params.filename

        if not file_path.exists():
            return json.dumps({
                "error": f"File '{params.filename}' not found in {DESIGN_DIR}"
            }, ensure_ascii=False)

        # If user specified sheet_name, force all unknown images to go there
        all_images = _extract_images_from_file(file_path, IMAGE_OUTPUT_DIR, params.sheet_name)

        if params.sheet_name:
            all_images = [img for img in all_images if img["sheet_name"] == params.sheet_name]

        return json.dumps({
            "filename": params.filename,
            "total_extracted": len(all_images),
            "output_directory": str(IMAGE_OUTPUT_DIR / file_path.stem),
            "images": all_images
        }, indent=2, ensure_ascii=False)

    except Exception as e:
        return json.dumps({
            "error": f"Error extracting images: {str(e)}"
        }, ensure_ascii=False)


@mcp.tool(
    name="ccs_get_design_info",
    annotations={
        "title": "Get Design File Information",
        "readOnlyHint": True,
        "destructiveHint": False,
        "idempotentHint": True,
        "openWorldHint": False
    }
)
async def ccs_get_design_info(params: GetDesignInfoInput, ctx: Context) -> str:
    """Get basic information about a design file: list of worksheets, file size, etc.

    Use this to quickly see what sheets are available in a design file before reading them.

    Args:
        params: Validated input parameters containing:
            - filename (str): Filename of the Excel design document

    Returns:
        str: JSON formatted information including sheet list and file stats
    """
    try:
        file_path = DESIGN_DIR / params.filename

        if not file_path.exists():
            return json.dumps({
                "error": f"File '{params.filename}' not found in {DESIGN_DIR}"
            }, ensure_ascii=False)

        if not file_path.is_file():
            return json.dumps({
                "error": f"'{params.filename}' is not a file"
            }, ensure_ascii=False)

        wb = load_workbook(filename=file_path, read_only=True, data_only=True)
        sheet_names = wb.sheetnames
        wb.close()

        stats = {
            "size_kb": round(file_path.stat().st_size / 1024, 1),
            "sheet_count": len(sheet_names)
        }

        return json.dumps({
            "filename": params.filename,
            "path": str(file_path),
            "stats": stats,
            "sheets": sheet_names
        }, indent=2, ensure_ascii=False)

    except Exception as e:
        return json.dumps({
            "error": f"Error getting design info: {str(e)}"
        }, ensure_ascii=False)


@mcp.tool(
    name="ccs_get_output_paths",
    annotations={
        "title": "Get Output Directory Paths",
        "readOnlyHint": True,
        "destructiveHint": False,
        "idempotentHint": True,
        "openWorldHint": False
    }
)
async def ccs_get_output_paths(ctx: Context) -> str:
    """Get the absolute paths to output directories where extracted images and files are saved.

    Returns:
        str: Output directory paths
    """
    return json.dumps({
        "output_dir": str(OUTPUT_DIR.absolute()),
        "image_output_dir": str(IMAGE_OUTPUT_DIR.absolute())
    }, indent=2, ensure_ascii=False)


# Database operation input models
class ExecuteSqlInput(BaseModel):
    """Input model for executing SQL statements."""
    model_config = ConfigDict(
        str_strip_whitespace=True,
        validate_assignment=True,
        extra='forbid'
    )

    sql: str = Field(
        ...,
        description="SQL statement to execute (e.g., CREATE TABLE, DROP TABLE, ALTER TABLE)",
        min_length=1
    )


class DescribeTableInput(BaseModel):
    """Input model for describing a table structure."""
    model_config = ConfigDict(
        str_strip_whitespace=True,
        validate_assignment=True,
        extra='forbid'
    )

    table_name: str = Field(
        ...,
        description="Name of the table to describe",
        min_length=1
    )


# Database tool definitions
@mcp.tool(
    name="ccs_execute",
    annotations={
        "title": "Execute SQL Statement",
        "readOnlyHint": False,
        "destructiveHint": True,
        "idempotentHint": False,
        "openWorldHint": True,
    }
)
async def ccs_execute(params: ExecuteSqlInput, ctx: Context) -> str:
    """Execute a SQL statement on the PostgreSQL database (e.g., CREATE TABLE, DROP TABLE, ALTER TABLE).

    Use this to create database tables based on design specifications.

    Args:
        params: Validated input parameters containing:
            - sql (str): SQL statement to execute

    Returns:
        str: JSON formatted result with execution status
    """
    conn = None
    try:
        conn = _get_db_connection()
        cursor = conn.cursor()
        cursor.execute(params.sql)
        conn.commit()

        row_count = cursor.rowcount

        return json.dumps({
            "success": True,
            "rows_affected": row_count,
            "message": "SQL executed successfully"
        }, indent=2, ensure_ascii=False)

    except Exception as e:
        if conn:
            conn.rollback()
        return json.dumps({
            "error": f"SQL execution failed: {str(e)}",
            "success": False
        }, ensure_ascii=False)
    finally:
        if conn:
            conn.close()


@mcp.tool(
    name="ccs_list_tables",
    annotations={
        "title": "List All Tables in Database",
        "readOnlyHint": True,
        "destructiveHint": False,
        "idempotentHint": True,
        "openWorldHint": False,
    }
)
async def ccs_list_tables(ctx: Context) -> str:
    """List all user-created tables in the public schema of the database.

    Returns:
        str: JSON formatted list of tables with their basic information
    """
    conn = None
    try:
        conn = _get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)

        sql = """
        SELECT table_name,
               (SELECT pg_total_relation_size(quote_ident(table_name)) / 1024 / 1024) as size_mb
        FROM information_schema.tables
        WHERE table_schema = 'public'
        AND table_type = 'BASE TABLE'
        ORDER BY table_name;
        """
        cursor.execute(sql)
        tables = cursor.fetchall()

        return json.dumps({
            "database": DB_NAME,
            "total_tables": len(tables),
            "tables": tables
        }, indent=2, ensure_ascii=False)

    except Exception as e:
        return json.dumps({
            "error": f"Failed to list tables: {str(e)}",
            "success": False
        }, ensure_ascii=False)
    finally:
        if conn:
            conn.close()


@mcp.tool(
    name="ccs_describe_table",
    annotations={
        "title": "Describe Table Structure",
        "readOnlyHint": True,
        "destructiveHint": False,
        "idempotentHint": True,
        "openWorldHint": False,
    }
)
async def ccs_describe_table(params: DescribeTableInput, ctx: Context) -> str:
    """Get detailed structure information about a specific table.

    This includes column names, data types, constraints, nullability, etc.

    Args:
        params: Validated input parameters containing:
            - table_name (str): Name of the table to describe

    Returns:
        str: JSON formatted table structure information
    """
    conn = None
    try:
        conn = _get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)

        # Get column information
        sql = """
        SELECT column_name,
               data_type,
               character_maximum_length,
               is_nullable,
               column_default
        FROM information_schema.columns
        WHERE table_schema = 'public'
          AND table_name = %s
        ORDER BY ordinal_position;
        """
        cursor.execute(sql, (params.table_name,))
        columns = cursor.fetchall()

        # Get primary key information
        pk_sql = """
        SELECT kcu.column_name
        FROM information_schema.table_constraints tc
        JOIN information_schema.key_column_usage kcu
          ON tc.constraint_name = kcu.constraint_name
         AND tc.table_schema = kcu.table_schema
         AND tc.table_name = kcu.table_name
        WHERE tc.constraint_type = 'PRIMARY KEY'
          AND tc.table_schema = 'public'
          AND tc.table_name = %s;
        """
        cursor.execute(pk_sql, (params.table_name,))
        primary_keys = [row['column_name'] for row in cursor.fetchall()]

        if not columns:
            return json.dumps({
                "error": f"Table '{params.table_name}' does not exist in public schema",
                "success": False
            }, ensure_ascii=False)

        # Get table comment
        comment_sql = """
        SELECT obj_description(%s::regclass);
        """
        cursor.execute(comment_sql, (params.table_name,))
        comment = cursor.fetchone()[0]

        return json.dumps({
            "table_name": params.table_name,
            "success": True,
            "columns": columns,
            "primary_keys": primary_keys,
            "comment": comment
        }, indent=2, ensure_ascii=False)

    except Exception as e:
        return json.dumps({
            "error": f"Failed to describe table: {str(e)}",
            "success": False
        }, ensure_ascii=False)
    finally:
        if conn:
            conn.close()


if __name__ == "__main__":
    mcp.run()
