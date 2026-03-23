"""
Extract text content from all sheets in Excel files
"""

import os
import sys
from pathlib import Path
import openpyxl
from openpyxl.utils import get_column_letter

def extract_text_from_excel(excel_path: Path, output_dir: Path):
    """Extract all text content from Excel sheets"""
    print(f"Processing: {excel_path.name}")

    wb = openpyxl.load_workbook(excel_path, data_only=True)

    # Create output file for this Excel
    output_path = output_dir / f"{excel_path.stem}_text.md"

    content = []
    content.append(f"# {excel_path.name}\n\n")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        content.append(f"---\n\n## Sheet: {sheet_name}\n\n")

        # Collect all rows with data
        rows_with_data = []
        max_cols = 0

        for row in ws.iter_rows():
            row_values = []
            has_data = False
            for cell in row:
                val = cell.value
                if val is not None and val != '':
                    has_data = True
                if isinstance(val, (int, float)):
                    row_values.append(f"{val}")
                elif val is None:
                    row_values.append("")
                else:
                    # Escape pipe characters
                    val_str = str(val).replace('|', r'\|')
                    # Remove newlines
                    val_str = val_str.replace('\n', ' ')
                    row_values.append(val_str)

            if has_data:
                rows_with_data.append(row_values)
                if len(row_values) > max_cols:
                    max_cols = len(row_values)

        if max_cols == 0:
            continue

        # Build markdown table
        content.append("| " + " | ".join([""] * max_cols) + " |\n")
        content.append("| " + " | ".join(["---"] * max_cols) + " |\n")

        for row_values in rows_with_data:
            # Pad row to max_cols
            while len(row_values) < max_cols:
                row_values.append("")
            content.append(f"| {' | '.join(row_values)} |\n")

        content.append("\n")

    # Write to file
    with open(output_path, 'w', encoding='utf-8') as f:
        f.writelines(content)

    print(f"  Saved text to: {output_path.name}")
    return len(wb.sheetnames)

def main():
    # 设计书目录
    design_dir = Path(__file__).parent / "设计书"
    # 输出目录
    output_dir = Path(__file__).parent / "extracted_text"
    output_dir.mkdir(exist_ok=True)

    total_sheets = 0

    # 处理所有 xlsx 文件
    for excel_file in design_dir.glob("*.xlsx"):
        if excel_file.name.startswith("~$"):  # skip temp files
            continue
        sheets = extract_text_from_excel(excel_file, output_dir)
        total_sheets += sheets

    print(f"\n=== Complete ===")
    print(f"Total {total_sheets} sheets processed")
    print(f"Text saved to: {output_dir}")

if __name__ == "__main__":
    main()
